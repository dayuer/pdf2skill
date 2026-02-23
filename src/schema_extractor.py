"""L1 骨架提取器 — 从保障利益表 Excel 中提取结构化产品骨架。

输出 skeleton.json：
{
  "product_name": "...",
  "responsibilities": [
    { "name": "住院医疗", "limit": "200万", "deductible": "1万", "ratio": "100%/60%", ... }
  ],
  "plans": [...],  // 保障计划（如有多档）
  "source_file": "保障利益表.xlsx"
}
"""

from __future__ import annotations

import json
import logging
from pathlib import Path

import openpyxl

_log = logging.getLogger(__name__)


def extract_skeleton(excel_path: str | Path) -> dict:
    """从保障利益表 Excel 提取结构化骨架。

    策略：
    1. 读取所有 sheet
    2. 识别表头行（含"保障责任"/"责任"/"项目"等关键词的行）
    3. 按表头映射列，逐行提取责任项
    """
    excel_path = Path(excel_path)
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    skeleton: dict = {
        "product_name": "",
        "responsibilities": [],
        "plans": [],
        "raw_sheets": {},
        "source_file": excel_path.name,
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = _read_sheet_rows(ws)
        if not rows:
            continue

        # 存原始数据（供 LLM 后续精炼）
        skeleton["raw_sheets"][sheet_name] = rows

        # 尝试识别表头并提取结构化数据
        header_idx, col_map = _find_header(rows)
        if header_idx is not None and col_map:
            items = _extract_items(rows, header_idx, col_map)
            skeleton["responsibilities"].extend(items)

            # 从表头上方寻找产品名称
            if not skeleton["product_name"]:
                skeleton["product_name"] = _find_product_name(rows, header_idx)

    wb.close()
    _log.info("骨架提取完成: %s, %d 项责任", excel_path.name, len(skeleton["responsibilities"]))
    return skeleton


def save_skeleton(skeleton: dict, output_path: str | Path) -> Path:
    """将骨架保存为 JSON。"""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(skeleton, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return output_path


# ── 内部工具 ──

# 表头关键词 → 标准列名
_HEADER_KEYWORDS = {
    "保障责任": "name",
    "责任名称": "name",
    "保障项目": "name",
    "项目": "name",
    "保险金额": "limit",
    "保额": "limit",
    "限额": "limit",
    "年限额": "limit",
    "免赔额": "deductible",
    "免赔": "deductible",
    "赔付比例": "ratio",
    "给付比例": "ratio",
    "报销比例": "ratio",
    "等待期": "waiting_period",
    "保障范围": "scope",
    "说明": "note",
    "备注": "note",
}


def _read_sheet_rows(ws) -> list[list[str]]:
    """读取 sheet 所有行，每个 cell 转为字符串。"""
    rows = []
    for row in ws.iter_rows():
        cells = [str(cell.value).strip() if cell.value is not None else "" for cell in row]
        # 跳过全空行
        if any(c for c in cells):
            rows.append(cells)
    return rows


def _find_header(rows: list[list[str]]) -> tuple[int | None, dict[int, str]]:
    """在前 10 行中找到表头行，返回 (行索引, {列索引: 标准列名})。"""
    for i, row in enumerate(rows[:10]):
        col_map: dict[int, str] = {}
        for j, cell in enumerate(row):
            for kw, std_name in _HEADER_KEYWORDS.items():
                if kw in cell:
                    col_map[j] = std_name
                    break
        # 至少匹配到 "责任名称" 才算有效表头
        if "name" in col_map.values():
            return i, col_map
    return None, {}


def _extract_items(
    rows: list[list[str]],
    header_idx: int,
    col_map: dict[int, str],
) -> list[dict]:
    """从表头下方逐行提取责任项。"""
    items = []
    for row in rows[header_idx + 1:]:
        item: dict[str, str] = {}
        for col_idx, field_name in col_map.items():
            if col_idx < len(row):
                val = row[col_idx]
                if val and val != "None":
                    item[field_name] = val
        # 至少有 name 才算有效
        if item.get("name"):
            items.append(item)
    return items


def _find_product_name(rows: list[list[str]], header_idx: int) -> str:
    """在表头上方寻找产品名称（通常是第一行的合并单元格）。"""
    for i in range(header_idx):
        for cell in rows[i]:
            if len(cell) > 5 and ("保险" in cell or "产品" in cell):
                return cell
    return ""
