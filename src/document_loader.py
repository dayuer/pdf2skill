"""
多格式文档加载器 — 支持 PDF、TXT、EPUB、DOCX、Excel。

职责：将不同格式的文档统一转换为 Markdown 文本，交给下游 Chunker。

格式处理策略：
- PDF：调用 MinerU (magic-pdf) 进行版面分析和结构化转换
- TXT：直接读取，按空行分段
- EPUB：通过 ebooklib 提取 HTML 章节，转为 Markdown
- DOCX：通过 python-docx 提取段落/表格/标题
- Excel：通过 openpyxl 将每个 sheet 转为 Markdown 表格
"""

from __future__ import annotations

import html
import re
import subprocess
import tempfile
from dataclasses import dataclass
from enum import Enum
from pathlib import Path
from typing import Optional


class DocFormat(Enum):
    """支持的文档格式"""

    PDF = "pdf"
    TXT = "txt"
    EPUB = "epub"
    DOCX = "docx"
    EXCEL = "excel"


@dataclass
class LoadResult:
    """文档加载结果"""

    markdown: str
    doc_name: str
    format: DocFormat
    # 原始文件路径
    source_path: str
    # 提取的目录树（如有）
    toc: Optional[list[str]] = None
    # 警告信息（如解析质量问题）
    warnings: list[str] | None = None


# ──── 格式检测 ────


def detect_format(filepath: str | Path) -> DocFormat:
    """根据文件扩展名检测文档格式"""
    suffix = Path(filepath).suffix.lower()
    mapping = {
        ".pdf": DocFormat.PDF,
        ".txt": DocFormat.TXT,
        ".text": DocFormat.TXT,
        ".md": DocFormat.TXT,
        ".markdown": DocFormat.TXT,
        ".epub": DocFormat.EPUB,
        ".docx": DocFormat.DOCX,
        ".doc": DocFormat.DOCX,
        ".xlsx": DocFormat.EXCEL,
        ".xls": DocFormat.EXCEL,
        ".csv": DocFormat.EXCEL,
    }
    fmt = mapping.get(suffix)
    if fmt is None:
        raise ValueError(f"不支持的文件格式：{suffix}（支持：PDF, TXT, EPUB, DOCX, Excel）")
    return fmt


# ──── TXT 加载器 ────


def _load_txt(filepath: Path) -> LoadResult:
    """加载纯文本 / Markdown 文件"""
    text = filepath.read_text(encoding="utf-8")
    return LoadResult(
        markdown=text,
        doc_name=filepath.stem,
        format=DocFormat.TXT,
        source_path=str(filepath),
    )


# ──── EPUB 加载器 ────

# HTML 标签 → Markdown 转换规则
_HTML_HEADING_RE = re.compile(r"<h([1-6])[^>]*>(.*?)</h\1>", re.IGNORECASE | re.DOTALL)
_HTML_P_RE = re.compile(r"<p[^>]*>(.*?)</p>", re.IGNORECASE | re.DOTALL)
_HTML_LI_RE = re.compile(r"<li[^>]*>(.*?)</li>", re.IGNORECASE | re.DOTALL)
_HTML_BR_RE = re.compile(r"<br\s*/?>", re.IGNORECASE)
_HTML_STRONG_RE = re.compile(r"<(?:strong|b)[^>]*>(.*?)</(?:strong|b)>", re.IGNORECASE | re.DOTALL)
_HTML_EM_RE = re.compile(r"<(?:em|i)[^>]*>(.*?)</(?:em|i)>", re.IGNORECASE | re.DOTALL)
_HTML_TAG_RE = re.compile(r"<[^>]+>")


def _html_to_markdown(raw_html: str) -> str:
    """将简单 HTML 转为 Markdown（轻量级，不依赖外部库）"""
    text = raw_html

    # 处理标题
    def _heading_replace(m: re.Match) -> str:
        level = int(m.group(1))
        title = _strip_tags(m.group(2)).strip()
        return f"\n{'#' * level} {title}\n"

    text = _HTML_HEADING_RE.sub(_heading_replace, text)

    # 处理加粗和斜体
    text = _HTML_STRONG_RE.sub(r"**\1**", text)
    text = _HTML_EM_RE.sub(r"*\1*", text)

    # 处理段落
    text = _HTML_P_RE.sub(lambda m: f"\n\n{_strip_tags(m.group(1)).strip()}\n\n", text)

    # 处理列表项
    text = _HTML_LI_RE.sub(lambda m: f"\n- {_strip_tags(m.group(1)).strip()}", text)

    # 处理换行
    text = _HTML_BR_RE.sub("\n", text)

    # 清除剩余 HTML 标签
    text = _HTML_TAG_RE.sub("", text)

    # 解码 HTML 实体
    text = html.unescape(text)

    # 压缩连续空行
    text = re.sub(r"\n{3,}", "\n\n", text)

    return text.strip()


def _strip_tags(s: str) -> str:
    """移除所有 HTML 标签"""
    return _HTML_TAG_RE.sub("", s)


def _load_epub(filepath: Path) -> LoadResult:
    """
    加载 EPUB 文件，提取所有章节转为 Markdown。

    使用 ebooklib 读取 EPUB 的 spine（阅读顺序），
    逐章提取 HTML 内容并转为 Markdown。
    """
    try:
        import ebooklib
        from ebooklib import epub
    except ImportError:
        raise ImportError("EPUB 支持需要安装 ebooklib：pip install ebooklib")

    book = epub.read_epub(str(filepath), options={"ignore_ncx": True})

    # 提取书名
    title_meta = book.get_metadata("DC", "title")
    doc_name = title_meta[0][0] if title_meta else filepath.stem

    # 提取目录
    toc_items: list[str] = []
    for item in book.toc:
        if isinstance(item, epub.Link):
            toc_items.append(item.title)
        elif isinstance(item, tuple) and len(item) >= 1:
            # 嵌套目录
            section = item[0]
            if isinstance(section, epub.Section):
                toc_items.append(section.title)

    # 按阅读顺序提取正文
    md_parts: list[str] = []

    for item in book.get_items_of_type(ebooklib.ITEM_DOCUMENT):
        raw_html = item.get_content().decode("utf-8", errors="replace")
        md = _html_to_markdown(raw_html)
        if md.strip():
            md_parts.append(md)

    markdown = "\n\n---\n\n".join(md_parts)

    return LoadResult(
        markdown=markdown,
        doc_name=doc_name,
        format=DocFormat.EPUB,
        source_path=str(filepath),
        toc=toc_items if toc_items else None,
    )


# ──── PDF 加载器 ────


def _load_pdf(filepath: Path) -> LoadResult:
    """
    加载 PDF 文件。

    优先使用 MinerU (magic-pdf) 进行版面分析。
    如果 MinerU 不可用，降级为简单的文本提取。
    """
    warnings: list[str] = []

    # 尝试 MinerU
    try:
        md = _load_pdf_mineru(filepath)
        return LoadResult(
            markdown=md,
            doc_name=filepath.stem,
            format=DocFormat.PDF,
            source_path=str(filepath),
            warnings=warnings or None,
        )
    except (FileNotFoundError, subprocess.CalledProcessError, OSError) as e:
        warnings.append(f"MinerU 不可用（{e}），降级为 pymupdf 提取")

    # 降级：pymupdf 纯文本提取
    try:
        md = _load_pdf_pymupdf(filepath)
        return LoadResult(
            markdown=md,
            doc_name=filepath.stem,
            format=DocFormat.PDF,
            source_path=str(filepath),
            warnings=warnings or None,
        )
    except ImportError:
        warnings.append("pymupdf 不可用，降级为 pdftotext 命令行工具")

    # 最终降级：pdftotext 命令行
    md = _load_pdf_pdftotext(filepath)
    return LoadResult(
        markdown=md,
        doc_name=filepath.stem,
        format=DocFormat.PDF,
        source_path=str(filepath),
        warnings=warnings or None,
    )


def _load_pdf_mineru(filepath: Path) -> str:
    """使用 MinerU (magic-pdf) 解析 PDF → Markdown"""
    with tempfile.TemporaryDirectory() as tmpdir:
        result = subprocess.run(
            ["magic-pdf", "-p", str(filepath), "-o", tmpdir, "-m", "auto"],
            capture_output=True,
            text=True,
            timeout=600,
        )
        if result.returncode != 0:
            raise subprocess.CalledProcessError(
                result.returncode, "magic-pdf", result.stderr
            )

        # MinerU 输出结构：{tmpdir}/{filename}/auto/{filename}.md
        stem = filepath.stem
        md_candidates = list(Path(tmpdir).rglob("*.md"))
        if not md_candidates:
            raise FileNotFoundError(f"MinerU 未生成 Markdown 文件：{tmpdir}")

        # 取最大的 .md 文件（通常是主内容）
        md_file = max(md_candidates, key=lambda f: f.stat().st_size)
        return md_file.read_text(encoding="utf-8")


def _load_pdf_pymupdf(filepath: Path) -> str:
    """使用 pymupdf 提取 PDF 纯文本（降级方案）"""
    try:
        import fitz  # pymupdf
    except ImportError:
        raise ImportError("pymupdf 未安装：pip install pymupdf")

    doc = fitz.open(str(filepath))
    pages: list[str] = []
    for page in doc:
        text = page.get_text("text")
        if text.strip():
            pages.append(text.strip())
    doc.close()

    return "\n\n---\n\n".join(pages)


def _load_pdf_pdftotext(filepath: Path) -> str:
    """使用 pdftotext 命令行工具提取文本（最终降级）"""
    result = subprocess.run(
        ["pdftotext", "-layout", str(filepath), "-"],
        capture_output=True,
        text=True,
        timeout=120,
    )
    if result.returncode != 0:
        raise RuntimeError(
            f"pdftotext 执行失败：{result.stderr}\n"
            "请安装 poppler：brew install poppler"
        )
    return result.stdout


# ──── 统一入口 ────


def load_document(filepath: str | Path) -> LoadResult:
    """
    加载文档并转换为 Markdown。

    自动检测文件格式，调用对应的加载器。

    Args:
        filepath: 文件路径

    Returns:
        LoadResult 包含 Markdown 文本和元信息

    Raises:
        FileNotFoundError: 文件不存在
        ValueError: 不支持的格式
    """
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"文件不存在：{path}")

    fmt = detect_format(path)

    loader_map = {
        DocFormat.PDF: _load_pdf,
        DocFormat.TXT: _load_txt,
        DocFormat.EPUB: _load_epub,
        DocFormat.DOCX: _load_docx,
        DocFormat.EXCEL: _load_excel,
    }

    return loader_map[fmt](path)


# ── DOCX 加载器 ──


def _load_docx(filepath: Path) -> LoadResult:
    """加载 Word (.docx) 文档，提取段落/表格/标题转为 Markdown。"""
    try:
        from docx import Document
        from docx.enum.text import WD_ALIGN_PARAGRAPH
    except ImportError:
        raise ImportError("DOCX 支持需要安装 python-docx：pip install python-docx")

    doc = Document(str(filepath))
    md_parts: list[str] = []

    for element in doc.element.body:
        tag = element.tag.split("}")[-1]  # 去掉命名空间

        if tag == "p":
            # 段落
            from docx.oxml.ns import qn
            style = element.find(qn("w:pPr"))
            style_name = ""
            if style is not None:
                pstyle = style.find(qn("w:pStyle"))
                if pstyle is not None:
                    style_name = pstyle.get(qn("w:val"), "")

            text = element.text or ""
            # 提取所有 run 的文本
            runs = element.findall(f".//{qn('w:t')}")
            if runs:
                text = "".join(r.text or "" for r in runs)

            if not text.strip():
                continue

            # 标题样式检测
            if style_name.startswith("Heading") or style_name.startswith("标题"):
                # 提取层级数字
                level = 1
                for ch in style_name:
                    if ch.isdigit():
                        level = int(ch)
                        break
                md_parts.append(f"{'#' * level} {text.strip()}")
            else:
                md_parts.append(text.strip())

        elif tag == "tbl":
            # 表格
            table = _extract_docx_table(element)
            if table:
                md_parts.append(table)

    return LoadResult(
        markdown="\n\n".join(md_parts),
        doc_name=filepath.stem,
        format=DocFormat.DOCX,
        source_path=str(filepath),
    )


def _extract_docx_table(tbl_element) -> str:
    """将 docx XML 表格元素转为 Markdown 表格。"""
    from docx.oxml.ns import qn

    rows_data: list[list[str]] = []
    for tr in tbl_element.findall(qn("w:tr")):
        cells = []
        for tc in tr.findall(qn("w:tc")):
            text_parts = [t.text or "" for t in tc.findall(f".//{qn('w:t')}")] 
            cells.append(" ".join(text_parts).strip())
        rows_data.append(cells)

    if not rows_data:
        return ""

    # 统一列数
    max_cols = max(len(r) for r in rows_data)
    for r in rows_data:
        while len(r) < max_cols:
            r.append("")

    # 生成 Markdown 表格
    lines = []
    lines.append("| " + " | ".join(rows_data[0]) + " |")
    lines.append("| " + " | ".join("---" for _ in rows_data[0]) + " |")
    for row in rows_data[1:]:
        lines.append("| " + " | ".join(row) + " |")

    return "\n".join(lines)


# ── Excel 加载器 ──


def _load_excel(filepath: Path) -> LoadResult:
    """加载 Excel (.xlsx/.xls/.csv)，将每个 sheet 转为 Markdown 表格。"""
    suffix = filepath.suffix.lower()

    if suffix == ".csv":
        return _load_csv(filepath)

    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("Excel 支持需要安装 openpyxl：pip install openpyxl")

    wb = load_workbook(str(filepath), read_only=True, data_only=True)
    md_parts: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        md_parts.append(f"## {sheet_name}")

        # 过滤全空行
        non_empty = [r for r in rows if any(c is not None for c in r)]
        if not non_empty:
            continue

        max_cols = max(len(r) for r in non_empty)

        def _cell(val) -> str:
            if val is None:
                return ""
            return str(val).replace("|", "│").replace("\n", " ").strip()

        # 第一行作为表头
        header = non_empty[0]
        header_cells = [_cell(c) for c in header] + [""] * (max_cols - len(header))
        lines = []
        lines.append("| " + " | ".join(header_cells) + " |")
        lines.append("| " + " | ".join("---" for _ in header_cells) + " |")

        for row in non_empty[1:]:
            cells = [_cell(c) for c in row] + [""] * (max_cols - len(row))
            lines.append("| " + " | ".join(cells) + " |")

        md_parts.append("\n".join(lines))

    wb.close()

    return LoadResult(
        markdown="\n\n".join(md_parts),
        doc_name=filepath.stem,
        format=DocFormat.EXCEL,
        source_path=str(filepath),
    )


def _load_csv(filepath: Path) -> LoadResult:
    """加载 CSV 文件转为 Markdown 表格。"""
    import csv

    with open(filepath, "r", encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f)
        rows = list(reader)

    if not rows:
        return LoadResult(
            markdown="",
            doc_name=filepath.stem,
            format=DocFormat.EXCEL,
            source_path=str(filepath),
        )

    def _cell(val: str) -> str:
        return val.replace("|", "│").replace("\n", " ").strip()

    lines = []
    lines.append("| " + " | ".join(_cell(c) for c in rows[0]) + " |")
    lines.append("| " + " | ".join("---" for _ in rows[0]) + " |")
    for row in rows[1:]:
        # 补齐列数
        padded = row + [""] * (len(rows[0]) - len(row))
        lines.append("| " + " | ".join(_cell(c) for c in padded) + " |")

    return LoadResult(
        markdown="\n".join(lines),
        doc_name=filepath.stem,
        format=DocFormat.EXCEL,
        source_path=str(filepath),
    )
